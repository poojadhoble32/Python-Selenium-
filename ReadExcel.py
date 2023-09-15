import openpyxl

def readexcel():
    file = "D:\Mine\python\SeleniumProjects\DemoFiles\Marvellousread.xlsx"

    # workbook object is created
    workbookobj = openpyxl.load_workbook(file)

    sheetobj = workbookobj.active
    colsize = sheetobj.max_column
    rowsize = sheetobj.max_row
    print(rowsize, colsize)

    # Loop will print all columns name
    for rowno in range(1, rowsize+1):
        for colno in range(1, colsize+1):
            cell = sheetobj.cell(row = rowno, column = colno)
            print(cell.value, end=" ")
        print()


if __name__ == '__main__':
    readexcel()